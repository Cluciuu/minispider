#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# ===============================================================================
#
# Copyright (c) 2021 Baidu.com, Inc. All Rights Reserved
#
# ===============================================================================
"""
模块介绍：

Authors: zhangkang06(zhangkang06@baidu.com)
Date:2022/03/25 23:55:20
"""

import os
import time
import random
import requests
import openpyxl
import datetime
import pytz
import hashlib
import threading
import urllib3
from lxml import etree
from openpyxl import utils
from UtilLib.osutil import OsUtil
from UtilLib.osutil import logutil
from UtilLib import threadsutil
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from requests.packages.urllib3.exceptions import InsecureRequestWarning


# lilla.liang@mdpi.com
# Md81b0c41ee37b

class MdpiSpider:
    TAG = 'MdpiSpider'
    logger = logutil.get_logger(TAG, need_write=False)

    def __init__(self, excel_file, password):
        self.headers = {
            "user-agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Mobile Safari/537.36"
        }
        self.proxy = None
        self.password = password
        self.session = requests.session()
        self.stroage_directory = self.creat_stroage_directory()
        self.timezone = self.get_time_zone()
        self.scholar_info_ls = None
        self.check_email_is_proposed_dict = None
        self.add_email_count_dict = None
        self.success_set = set()
        self.excel_file = excel_file
        self.excel_src_md5 = None

    def creat_stroage_directory(self):
        directory = os.path.join(os.path.expanduser('~/Desktop'), 'MdpiSpiderProduct')
        OsUtil.creat_directory(directory)
        return directory

    def start_work(self):
        self.login()
        self.sleep_random_time()
        self.uniform_excel(self.excel_file)
        available_issue_list = self.get_available_issue_list()
        self.sleep_random_time()

        self.scholar_info_ls = self.get_scholar_ls_from_excel(self.excel_file)
        self.scholar_info_ls.sort(key=lambda x: x[1])
        threadsutil.wait_time_async_run(0.5, fun=self.excel_is_changed, args=(self.excel_file,))

        issue_url, issue_id = self.get_an_issue(available_issue_list)
        self.add_email_to_issue(issue_url, issue_id)

    def login(self):
        MdpiSpider.logger.info(f'start logging')
        url = 'https://login.mdpi.com'
        requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
        response = self.session.get(url=url)
        MdpiSpider.logger.info(f'response:{response}')
        selector = etree.HTML(response.text)
        csrf_token = selector.xpath('//input[@name="_csrf_token"]/@value')[0]
        MdpiSpider.logger.info(f'csrf_token: {csrf_token}')
        post_data = {"_username": "lilla.liang@mdpi.com",
                     "_password": self.password,
                     "_target_path": "https://login.mdpi.com/",
                     "_csrf_token": csrf_token

                     }
        url = 'https://login.mdpi.com/login_check'
        # url = response.url
        response = self.session.post(url=url, data=post_data, verify=False)
        MdpiSpider.logger.info(f'login response.status_code: {response.status_code}')

        file_name = os.path.join(self.stroage_directory, 'AfterLoginMdpi.html')
        # self.write_to_file(response.text, file_name)
        if 'forgot' not in response.text:
            MdpiSpider.logger.info(f'login status: success')
        else:
            MdpiSpider.logger.info(f'Incorrect password')
            raise ConnectionRefusedError

    def get_available_issue_list(self):
        url = 'https://susy.mdpi.com/special_issue_pending/list'
        response = self.session.get(url=url, timeout=3)
        file_name = os.path.join(self.stroage_directory, 'SpecialIssues.html')
        # self.write_to_file(response.text, file_name)--
        selector = etree.HTML(response.text)
        all_issue_ls = selector.xpath('//tr[@data-id]')
        MdpiSpider.logger.debug(f'issue_ls : {all_issue_ls}')
        available_issue_list = []
        for issue_item in all_issue_ls:
            # issue id
            issue_item_id = issue_item.get('data-id')
            issue_item_children = list(issue_item)
            # issue名称 = issue_item_children[1]
            issue_name_ls = list(issue_item_children[1])
            issue_name = issue_name_ls[0].text
            # issue状态 = issue_item_children[5]
            MdpiSpider.logger.debug(f'{issue_item_id:7}【 {issue_name:70} 】｜ status: {issue_item_children[5].text}')
            if 'Pending GE invitation' in issue_item_children[5].text:
                available_issue_list.append((issue_item_id, issue_name))
        return available_issue_list

    def get_an_issue(self, available_issue_list):
        issue_id = available_issue_list[0][0]
        issue_name = available_issue_list[0][1]
        issue_url = f'https://susy.mdpi.com/special_issue/process/{issue_id}'
        MdpiSpider.logger.info(
            f'Issue to be added: [ issue_id：{issue_id}, issue_name: {issue_name}, issue_url:{issue_url} ]')
        response = self.session.get(url=issue_url, timeout=3)
        issue_file_name = os.path.join(self.stroage_directory, f'{issue_id}.html')
        # self.write_to_file(response.text, issue_file_name)
        return issue_url, issue_id

    def add_email_to_issue(self, issue_url, issue_id):
        MdpiSpider.logger.info('开始检测&&邀请')
        stop = False
        thread_ls = []
        # msg_queue = Queue()
        self.check_email_is_proposed_dict = {}
        self.add_email_count_dict = {}
        while not stop:
            try:
                paris_strft_time_now, paris_strpt_time_now = self.get_paris_time()
                for scholar_info in self.scholar_info_ls:
                    email_address = scholar_info[0]
                    scholar_name = email_address.split('@')[0]
                    last_operation_strft_time = scholar_info[1]
                    last_operation_strpt_time = datetime.datetime.strptime(last_operation_strft_time,
                                                                           "%Y-%m-%d %H:%M:%S")
                    scholar_status = scholar_info[2]
                    handle_status = scholar_info[3]
                    timedelta = datetime.timedelta(days=90) - (paris_strpt_time_now - last_operation_strpt_time)
                    if datetime.timedelta(hours=-24) < timedelta <= datetime.timedelta(seconds=30) \
                            and email_address not in thread_ls \
                            and handle_status is None:
                        if datetime.timedelta(hours=-24) < timedelta < datetime.timedelta(seconds=0):
                            MdpiSpider.logger.info(f'{email_address} | 已到达可邀请时间，捡漏尝试')
                        if datetime.timedelta(seconds=0) < timedelta <= datetime.timedelta(seconds=30):
                            MdpiSpider.logger.info(f'{email_address} | 即将到达可邀请时间，准备开始邀请')
                        self.check_email_is_proposed_dict[email_address] = None
                        self.add_email_count_dict[email_address] = 0
                        form_token = self.get_form_token(email_address, issue_url)
                        if form_token is None:
                            continue
                        else:
                            self.start_add_email(scholar_name, email_address, form_token, issue_url, scholar_info,
                                                 issue_id)
                            thread_ls.append(email_address)
                    if len(self.success_set) > 20:
                        stop = True
                time.sleep(5)
            except:
                MdpiSpider.logger.exception('程序运行出错')

    def start_add_email(self, scholar_name, email_address, form_token, issue_url, scholar_info, issue_id):
        add_thread_ls = [
            threading.Thread(target=self.add_an_email_to_issue, name=f'{scholar_name}_thread_{i}',
                             args=(email_address, form_token, issue_url, scholar_info,)) for i in range(4)]
        for thread in add_thread_ls:
            if __name__ == '__main__':
                thread.start()
                time.sleep(0.1)

        check_thread = threading.Thread(target=self.check_email_is_proposed, args=(email_address, issue_url, issue_id,))
        check_thread.start()

    def get_form_token(self, email, issue_url):
        MdpiSpider.logger.info(f"target_issue_url:{issue_url}")
        form_token = None
        err_count = 0
        view_response = None
        while not form_token:
            try:
                view_response = self.session.get(url=issue_url)
            except:
                MdpiSpider.logger.exception(f'session get error: {issue_url}')
            if view_response and email in view_response.text:
                MdpiSpider.logger.info(
                    f'{threading.current_thread().name} | [{email}] has already existed in this issue: {issue_url}')
                self.modify_email_status(email, '已存在')
                return None
            try:
                selector = etree.HTML(view_response.text)
                form_token = selector.xpath('//*[@id="form__token"]/@value')[0]
                MdpiSpider.logger.info(f'form_token: {form_token}')
            except:
                MdpiSpider.logger.info(f'{email} form_token 获取错误，尝试重新登录之后再获取')
                err_count += 1
                self.login()
                if err_count > 3:
                    MdpiSpider.logger.info(f'因不能获取from_token，提前结束本次操作:{email}，下轮循环中依然有机会抢到')
                    break
        return form_token

    def add_an_email_to_issue(self, email, form_token, issue_url, scholar_info):
        post_data = {
            "form[email]": email,
            "form[_token]": form_token
        }
        stop = False
        while not stop:
            start_time = time.time()
            try:
                self.session.post(url=issue_url, data=post_data)
            except:
                MdpiSpider.logger.exception(f'session post error: {issue_url}')
            stop = self.check_email_is_proposed_dict.get(email, None)
            end_time = time.time()
            MdpiSpider.logger.info(f'check_email_is_proposed: {self.check_email_is_proposed_dict}')
            self.add_email_count_dict[email] = self.add_email_count_dict[email] + 1
            timedelta = self.get_real_timedelta(scholar_info)
            MdpiSpider.logger.info(
                f'{threading.current_thread().name} ｜ {email} | time left: {timedelta} | 第 {self.add_email_count_dict[email]} 次尝试 | cost time: {end_time - start_time}')

    def check_email_is_proposed(self, email, issue_url, issue_id):
        scholar_name = email.split('@')[0]
        while self.check_email_is_proposed_dict[email] is not True:
            check_url = f'https://susy.mdpi.com/user/guest_editor/check?email={email}&special_issue_id={issue_id}'
            check_result = self.session.get(check_url).text

            if 'This editor was proposed as GE by' in check_result:
                paris_time = self.get_paris_time()[0]
                cn_time = self.get_cn_time()
                self.check_email_is_proposed_dict[email] = True
                selector = etree.HTML(check_result)
                text_ls = selector.xpath('//p/text()')
                name = None
                if text_ls:
                    text = text_ls[0]
                    pretext = text.split('on')
                    name = pretext[0].split('by')[-1]
                    MdpiSpider.logger.info(f'{scholar_name} | was invited by {name}')
                if name is not None and 'Lilla Liang' in name:
                    MdpiSpider.logger.info(f'{scholar_name} added successfully: {issue_url}')
                    self.success_set.add(email)
                    MdpiSpider.logger.info(f'本次共成功邀请 {len(self.success_set)} 个: {self.success_set}')
                    self.modify_email_status(email, f'已添加 \nparis time:\n{paris_time} \ncn time:\n{cn_time}')
                else:
                    self.modify_email_status(email, f'被 {name} 添加 \nparis time:\n{paris_time} \ncn time:\n{cn_time}')

            elif 'The number of proposed GE cannot exceed 5 at most in each special issue' in check_result:
                self.check_email_is_proposed_dict[email] = True
                paris_time = self.get_paris_time()[0]
                cn_time = self.get_cn_time()
                MdpiSpider.logger.info(f'{scholar_name} added successfully: {issue_url}')
                self.success_set.add(email)
                MdpiSpider.logger.info(f'本次共成功邀请 {len(self.success_set)} 个: {self.success_set}')
                self.modify_email_status(email, f'已添加 \nparis time:\n{paris_time} \ncn time:\n{cn_time}')

            elif 'The editor was invited as Guest Editor within 90 days.' in check_result:
                MdpiSpider.logger.info(f'{scholar_name} | was invited as Guest Editor within 90 days')
            time.sleep(3)

    def modify_email_status(self, email, status):
        sheet = None
        excel = None
        try:
            excel = openpyxl.load_workbook(self.excel_file)
            sheet = excel.worksheets[0]
        except:
            MdpiSpider.logger.error('excel load error')
        if sheet is not None and excel is not None:
            row_index = 1
            for row in sheet.iter_rows():
                if row[0].value is not None and row[1].value is not None and row[2].value is not None:
                    email_address = row[0].value
                    # last_operation_strft_time = row[1].value
                    # last_operation_strpt_time = datetime.datetime.strptime(last_operation_strft_time, "%Y-%m-%d %H:%M:%S")
                    # scholar_status = row[2].value
                    if email_address == email:
                        save_sucess = False
                        while not save_sucess:
                            try:
                                sheet.cell(row_index, 5).value = status
                                excel.save(self.excel_file)
                                save_sucess = True
                                MdpiSpider.logger.info('excel save successfully')
                            except:
                                MdpiSpider.logger.info('excel save error')
                row_index += 1

    def get_real_timedelta(self, scholar_info):
        paris_strft_time_now, paris_strpt_time_now = self.get_paris_time()
        # email_address = scholar_info[0]
        last_operation_strft_time = scholar_info[1]
        last_operation_strpt_time = datetime.datetime.strptime(last_operation_strft_time, "%Y-%m-%d %H:%M:%S")
        # scholar_status = scholar_info[2]
        # handle_status = scholar_info[3]
        timedelta = datetime.timedelta(days=90) - (paris_strpt_time_now - last_operation_strpt_time)
        return timedelta

    def excel_is_changed(self, excel_file):
        while True:
            changed = False
            with open(excel_file, 'rb') as file:
                content = file.read()
            new_md5 = hashlib.md5(content).hexdigest()
            if new_md5 != self.excel_src_md5:
                MdpiSpider.logger.info('检测到文件变动，重新获取 scholar_info' + '\n' + '*' * 150)
                MdpiSpider.logger.info(f'orignal_md5: {self.excel_src_md5}, new_md5: {new_md5}')
                self.excel_src_md5 = new_md5
                self.scholar_info_ls = self.get_scholar_ls_from_excel(excel_file)
                info_str = "\n".join(str(info) for info in self.scholar_info_ls)
                MdpiSpider.logger.info(
                    f'scholar_info_ls changed: 共{len(self.scholar_info_ls)}个:\n{"=" * 150}\n{info_str}\n{"=" * 150}')
                changed = True
            if changed and len(self.scholar_info_ls) != 0:
                try:
                    last_operation_strft_time = self.scholar_info_ls[0][1]
                    last_operation_strpt_time = datetime.datetime.strptime(last_operation_strft_time,
                                                                           "%Y-%m-%d %H:%M:%S")
                    paris_strft_time_now, paris_strpt_time_now = self.get_paris_time()
                    timedelta = datetime.timedelta(days=90) - (paris_strpt_time_now - last_operation_strpt_time)
                    MdpiSpider.logger.info(f'sorted [0] name: {self.scholar_info_ls[0][0]}')
                    MdpiSpider.logger.info(f'sorted [0] last_operation_strft_time: {last_operation_strft_time}')
                    MdpiSpider.logger.info(f'sorted [0] timedelta: {timedelta}')
                    MdpiSpider.logger.info(f"本次共成功邀请 {len(self.success_set)} 个: {self.success_set}\n{'*' * 150}\n\n\n")
                except:
                    MdpiSpider.logger.info('scholar_info_ls 解析出错')
            time.sleep(1)

    def get_scholar_ls_from_excel(self, excel_file):
        excel = None
        sheet = None
        while not sheet or not excel:
            try:
                excel = openpyxl.load_workbook(excel_file)
                sheet = excel.worksheets[0]
            except:
                MdpiSpider.logger.error('excel read error')
        scholar_info_ls = []
        row_index = 1
        for row in sheet.iter_rows():
            MdpiSpider.logger.debug('=' * 150)
            if row[0].value is not None and row[1].value is not None and row[2].value is not None:
                email_address = row[0].value
                last_operation_strft_time = row[1].value
                last_operation_strpt_time = datetime.datetime.strptime(last_operation_strft_time, "%Y-%m-%d %H:%M:%S")
                scholar_status = row[2].value
                handle_status = sheet.cell(row_index, 5).value
                MdpiSpider.logger.debug(
                    f'{row_index} [email_address:{email_address:15} | last_operation_time:{last_operation_strft_time}]'
                    f'| status:{scholar_status} |handle_status:{handle_status} ]')
                paris_strft_time_now, paris_strpt_time_now = self.get_paris_time()
                timedelta = datetime.timedelta(days=90) - (paris_strpt_time_now - last_operation_strpt_time)
                cn_time_now = self.get_cn_time()
                if datetime.timedelta(hours=-5) < timedelta < datetime.timedelta(hours=24) and \
                        email_address is not None:
                    scholar_info_ls.append((email_address, last_operation_strft_time, scholar_status, handle_status,
                                            (f'刷新时间:{cn_time_now}',
                                             f'剩余时间:{timedelta}' if timedelta > datetime.timedelta(
                                                 days=0) else f'已经过了:{-timedelta}')))
            row_index += 1
        scholar_info_ls.sort(key=lambda x: x[1])
        return scholar_info_ls

    def write_to_file(self, text, file_name):
        threadsutil.wait_time_async_run(0.5, self.__write_to_file, args=(text, file_name,))

    def __write_to_file(self, text, file_name):
        MdpiSpider.logger.info(f'file_name: {file_name}')
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write(text)

    def uniform_excel(self, excel_file):
        MdpiSpider.logger.info('start uniforming excel')
        uniform_succeed = False
        while not uniform_succeed:
            try:
                excel = None
                sheet = None
                while not sheet or not excel:
                    try:
                        excel = openpyxl.load_workbook(excel_file)
                        sheet = excel.worksheets[0]
                    except:
                        MdpiSpider.logger.error('excel read error')
                width = 30
                for i in range(1, sheet.max_column + 1):
                    sheet.column_dimensions[utils.get_column_letter(i)].width = width

                # 初始化单元格对齐方式的对象
                alignment = Alignment(
                    horizontal='left',  # 水平对齐方式:center, left, right
                    vertical='center',  # 垂直对齐方式: center, top, bottom
                    wrapText=True  # 自动换行
                )

                for row in excel.active.rows:
                    for cell in row:
                        cell.alignment = alignment
                MdpiSpider.logger.info('excel uniformed successfully')
                excel.save(excel_file)
                uniform_succeed = True
            except:
                MdpiSpider.logger.info('failure to uniform excel')

    def sleep_random_time(self):
        random_seconds = random.uniform(0.1, 0.3)
        time.sleep(random_seconds)

    def transform_time_to_int(self, time):
        time_ls = time.split()
        pre_str = time_ls[0].replace('-', '')
        post_str = time_ls[1].replace(':', '')
        time_int = int(pre_str + post_str)
        return time_int

    def get_paris_time(self):
        paris_strft_time_now = datetime.datetime.now(self.timezone).strftime('%Y-%m-%d %H:%M:%S')
        paris_strpt_time_now = datetime.datetime.strptime(paris_strft_time_now, "%Y-%m-%d %H:%M:%S")
        return paris_strft_time_now, paris_strpt_time_now

    def get_cn_time(self):
        cn_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return cn_time

    def get_time_zone(self):
        timezone = pytz.timezone('Europe/Paris')
        MdpiSpider.logger.info(f'timezone: {timezone}')
        return timezone
